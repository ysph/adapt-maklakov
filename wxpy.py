import cv2
import numpy as np
from openpyxl.workbook.workbook import Workbook
# GUI library
import wx
import wx.grid

# excel
from openpyxl import load_workbook
from openpyxl.styles import Font
# custom detection algo
import detection as det
import calculate as calc

import array
import os

path = 'db.xlsx'
if os.path.exists(path):
    wb = load_workbook(path)
else:
    wb = Workbook()
    wb._sheets[0]['A1'] = 'Номер'
    wb._sheets[0]['B1'] = 'ФИО'
    wb._sheets[0]['C1'] = 'Подразделение'
    wb._sheets[0]['D1'] = 'Достоверность (Д)'
    wb._sheets[0]['E1'] = 'Нервно-психическая устойчивость (НПУ)'
    wb._sheets[0]['F1'] = 'Личностный адаптивный потенциал (АС)'
    wb._sheets[0]['G1'] = 'Коммуникативный потенциал (КП)'
    wb._sheets[0]['H1'] = 'Моральная нормативность (МН)'
    wb._sheets[0]['I1'] = 'Группа НПУ'

    wb._sheets[0].column_dimensions['B'].width = 58
    wb._sheets[0].column_dimensions['C'].width = 21
    wb._sheets[0].column_dimensions['D'].width = 40
    wb._sheets[0].column_dimensions['E'].width = 40
    wb._sheets[0].column_dimensions['F'].width = 40
    wb._sheets[0].column_dimensions['G'].width = 40
    wb._sheets[0].column_dimensions['H'].width = 40
    wb._sheets[0].column_dimensions['H'].width = 20
ws = wb._sheets[0]

def GetData():
    data = {}
    justdata = []
    for row in range(2,ws.max_row+1):
        fio = ws.cell(row=row,column=2).value
        unit = ws.cell(row=row,column=3).value
        if fio != '' and fio != None and unit != '' and unit != None:
            fio = fio.lower()
            unit = unit.lower()
            justdata.append(fio)
            if data.get(unit) == None:
                data[unit] = [fio]
            else:
                data[unit].append(fio)
    return data,justdata
data,justdata = GetData()

class MyCC(wx.TextCompleterSimple):
    def __init__(self,parent,tp,maxResults=25):
        wx.TextCompleterSimple.__init__(self)
        self.maxRes = maxResults
        self.type = tp
        self.parent = parent
    def GetCompletions(self,prefix):
        res = []
        prfx = prefix.lower()
        if self.type == 'unit':
            chooseFrom = list(data.keys())
        elif self.type == 'fio':
            unitWndw = self.parent.FindWindowByName('unitCtrl')
            unit = unitWndw.GetValue()
            if unit.strip() == '' or (unit not in data.keys()):
                chooseFrom = justdata
            elif unit in data.keys():
                chooseFrom = data[unit]
        for item in chooseFrom:
            if item.lower().startswith(prfx):
                res.append(item)
                if len(res) == self.maxRes:
                    return res
        return res

class Frame(wx.Frame):
    def calcResult(self):
        l = calc.testPsych(self.data)
        for i, res in enumerate(self.ress):
            res.SetLabel(str(l[i]))
        
    def GetOnMouseOver(self, grid, gridWdw):
        def OnMouseOver(event):
            #gridWdw.SetFocus()
            x,y = grid.CalcUnscrolledPosition(event.GetX(),event.GetY())
            coords = grid.XYToCell(x,y)
            row = coords[0]
            col = coords[1]
            if not(row == 8 and col >= 5):
                grid.SetGridCursor(row,col)
            event.Skip()
            
        return OnMouseOver
    def GetOnClick(self, grid):
        def OnClick(event):
            grid.ClearSelection()
            c = event.GetCol()
            r = event.GetRow()
            res = grid.GetCellValue(r,c)
            if '+' in res:
                grid.SetCellValue(r,c,res[:-1] + '–') # not minus
                self.data[(r*self.cols)+c+1] = '-'
                self.calcResult()
            elif '–' in res:
                grid.SetCellValue(r,c,res[:-1] + '+')
                self.data[(r*self.cols)+c+1] = '+'
                self.calcResult()
            grid.ForceRefresh()
        return OnClick
    def OnPress(self,event):
        codes = {388:'+',390:'–',61:'+',45:'–',314:'-1',315:'-1',316:'-1',317:'-1'}
        if event.GetKeyCode() not in codes:
            return
        r = self.grid.GetGridCursorCoords()[0]
        c = self.grid.GetGridCursorCoords()[1]
        res = self.grid.GetCellValue(r,c)
        code = codes[event.GetKeyCode()]
        
        if code not in res and code != '-1' and ('+' in res or '–' in res):
            self.grid.SetCellValue(r,c,res[:-1] + code) # not minus
            if code == '–':
                tmp = '-'
            else:
                tmp = '+'
            self.data[(r*self.cols)+c+1] = tmp # minus
            self.calcResult()
        self.grid.ForceRefresh()
        event.Skip()
    def __init__(self, parent, id):
        wx.Frame.__init__(self, parent, -1, 'Распознать', size=(1150, 650))
        self.panel = wx.Panel(self)
        self.width,self.height = 1160,200
        self.rows,self.cols = 9,20
        
        grid = self.CreateGrid()
        self.grid = grid
        
        #create empty image
        empty = wx.Image(self.width,self.height)
        a = array.array('B', empty.GetData())
        for i in range(len(a)):
            a[i] = 200
        empty.SetData(a.tobytes())
        empty = empty.ConvertToBitmap()
        self.mapq = wx.StaticBitmap(parent=self.panel, bitmap=empty, id=-1)
        
        # First create the controls
        topLbl = wx.StaticText(self.panel, -1, "«Адаптивность - 02» (МЛО-АМ)")
        topLbl.SetFont(wx.Font(18, wx.SWISS, wx.NORMAL, wx.BOLD))

        prev = wx.Image("icons/left_arrow.png", wx.BITMAP_TYPE_PNG)
        prev.Rescale(18,18)
        next = prev.Mirror()
        prevbtn = wx.BitmapButton(self.panel, -1, prev.ConvertToBitmap(),name='decrement_year')
        nextbtn = wx.BitmapButton(self.panel, -1, next.ConvertToBitmap(),name='increment_year')

        actionsSizer = wx.FlexGridSizer(rows=1,cols=6,vgap=0,hgap=5)
        openBtn = wx.Button(self.panel, -1, "Открыть")
        numberOfPeople = wx.StaticText(self.panel,-1,"из ")
        self.current_blank = wx.StaticText(self.panel,-1,"  ")
        self.number_blanks = wx.StaticText(self.panel,-1,"")
        self.IsLoaded = False
        actionsSizer.Add(openBtn,flag=wx.ALIGN_CENTER)
        actionsSizer.Add(prevbtn,flag=wx.ALIGN_CENTER)
        actionsSizer.Add(self.current_blank,flag=wx.ALIGN_CENTER)
        actionsSizer.Add(nextbtn,flag=wx.ALIGN_CENTER)
        actionsSizer.Add(numberOfPeople,flag=wx.ALIGN_CENTER)
        actionsSizer.Add(self.number_blanks,flag=wx.ALIGN_CENTER)

        # saving
        self.saveSizer = wx.FlexGridSizer(rows=2,cols=3,vgap=0,hgap=10)
        desc = "Подразделение,Фамилия Имя Отчество,Действие".split(',')
        self.unit = wx.TextCtrl(self.panel,-1,'', size=(100,30),name='unitCtrl')
        self.fio = wx.TextCtrl(self.panel, -1, "", size=(200,30),name='fioCtrl',style=wx.BORDER_SUNKEN)
        self.fio.Disable()
        self.saveBtn = wx.Button(self.panel, -1, "Сохранить")
        self.saveBtn.Disable()
        self.unit.AutoComplete(MyCC(self,'unit',maxResults=25))
        self.fio.AutoComplete(MyCC(self,'fio',maxResults=25))
        for d in desc:
            self.saveSizer.Add(wx.StaticText(self.panel,-1,d),flag=wx.ALIGN_CENTRE)
        self.saveSizer.Add(self.unit)
        self.saveSizer.Add(self.fio)
        self.saveSizer.Add(self.saveBtn)
        # result text
        resultSizer = wx.FlexGridSizer(rows=6, cols=2, hgap=45, vgap=0)
        self.labels = "Достоверность (Д),Нервно-психическая устойчивость (НПУ),Личностный адаптивный потенциал (АС),Коммуникативный потенциал (КП),Моральная нормативность (МН),Группа НПУ".split(',')
        self.ress = []
        for i,lbl in enumerate(self.labels):
            a = wx.StaticText(self.panel, wx.ID_ANY, lbl+':') #lbl+':'
            b = wx.StaticText(self.panel, wx.ID_ANY, '0', name=lbl)
            self.ress.append(b)
            resultSizer.Add(a,0,wx.ALIGN_LEFT)
            resultSizer.Add(b,0,wx.ALIGN_RIGHT)
        # mainSizer is the top-level one that manages everything
        self.mainSizer = wx.BoxSizer(wx.VERTICAL)
        self.mainSizer.Add(topLbl, 0, wx.CENTER, 5)
        self.mainSizer.Add(self.mapq, 0, wx.EXPAND|wx.ALL, 10)

        btnSizer = wx.BoxSizer(wx.HORIZONTAL)
        btnSizer.Add((0,20), 1)
        btnSizer.Add(actionsSizer)
        btnSizer.Add((0,20), 1)
        self.mainSizer.Add(btnSizer, 0, wx.EXPAND|wx.BOTTOM, 10)
        self.mainSizer.Add(grid,0,wx.CENTRE|wx.CENTRE)
        
        self.mainSizer.Add((0,10), 1)
        self.mainSizer.Add(resultSizer,0,wx.CENTRE|wx.CENTRE)
        self.mainSizer.Add((0,10), 1)
        self.mainSizer.Add(self.saveSizer,flag=wx.ALIGN_CENTER)
        self.mainSizer.Add((0,10))

        self.panel.SetSizer(self.mainSizer)
        self.mainSizer.Fit(self)
        self.mainSizer.SetSizeHints(self)
        
        # bind events
        self.Bind(wx.EVT_BUTTON, self.OnOpenImgs, openBtn)
        self.Bind(wx.EVT_BUTTON, self.OnNext, nextbtn)
        self.Bind(wx.EVT_BUTTON, self.OnPrev, prevbtn)
        self.Bind(wx.EVT_BUTTON, self.OnSave, self.saveBtn)
        self.Bind(wx.EVT_TEXT, self.OnPressUnit, self.unit)

        self.Bind(wx.grid.EVT_GRID_CELL_LEFT_CLICK, self.GetOnClick(grid))
        self.grid.Bind(wx.EVT_KEY_DOWN,self.OnPress)
        
        grid.GetGridWindow().Bind(wx.EVT_MOTION, self.GetOnMouseOver(grid, grid.GetGridWindow()))
    def OnPressUnit(self,event):
        if len(self.unit.GetValue()) <= 0:
            self.fio.Disable()
        else:
            self.fio.Enable()
    def OnOpenImgs(self, event):
        wcard = "JPEG (*.jpg,*.jpeg,*.jfif,*.jpe)|*.jpg|PNG (*.png)|*.png|TIFF (*.tif)|*.tif"
        dlg = wx.FileDialog(self, "Выбрать изображения", defaultDir="./", 
                            style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST | wx.FD_MULTIPLE, wildcard=wcard)
        if dlg.ShowModal() == wx.ID_OK:
            filenames = dlg.GetFilenames()
            filepaths = dlg.GetPaths()
            self.ReadNew(filepaths)
        dlg.Destroy()
    def ReadNew(self,paths):
        self.overallTables = {}
        for path in paths:
            for tbl in det.FindTables(path,self.rows,self.cols).values():
                self.overallTables[len(self.overallTables)+1] = tbl

        self.number_blanks.SetLabel(str(len(self.overallTables)))
        self.current_blank.SetLabel('1')
        self.IsLoaded = True
        self.saveBtn.Enable()
        self.LoadBlank()
    def nOfDs(self, number):
        new = str(number)
        if number < 10:
            return new + "    "
        elif number < 100:
            return new + "  "
        elif number < 1000:
            return new + " "
    def CreateGrid(self):
        grid = wx.grid.Grid(self.panel,style=wx.BORDER_SIMPLE)
        grid.CreateGrid(self.rows,self.cols,selmode=4)
        grid.HideColLabels()
        grid.HideRowLabels()
        
        font = wx.Font()
        font.SetFaceName("Times New Roman")
        font.SetPointSize(14)
        
        grid.DisableDragGridSize()
        for row in range(self.rows):
            for col in range(self.cols):
                v = ((row*self.cols) + col) + 1
                if v <= 165:
                    grid.SetCellValue(row, col, self.nOfDs(v) + ' ')
                grid.SetReadOnly(row,col)
        grid.SetDefaultRowSize(22, resizeExistingRows=True)
        grid.SetDefaultColSize(58, resizeExistingCols=True)
        grid.SetDefaultCellFont(font)
        grid.SetDefaultCellAlignment(wx.CENTRE,wx.CENTRE)
        grid.SetGridLineColour("#000000")
        return grid
    def LoadBlank(self):
        table = self.overallTables[self.GetCurrentPerson()]
        img = cv2.resize(table['image'], (self.width, self.height))
        bmp = wx.Bitmap.FromBuffer(img.shape[1],img.shape[0],img)
        self.mapq.SetBitmap(bmp)
        self.data = det.FindData(table['cells'])
            
        for k,v in self.data.items():
            c = (k - 1) % self.cols
            r = (k - c) // self.cols
            if v == '-':
                v = '–'
            txt = self.grid.GetCellValue(r,c)[:-1] + v
            self.grid.SetCellValue(r,c,txt)
        self.calcResult()
        self.grid.GetGridWindow().SetFocus()
        self.grid.ForceRefresh()
    def GetMaxPersons(self):
        return int(self.number_blanks.GetLabel())
    def GetCurrentPerson(self):
        return int(self.current_blank.GetLabel())
    def SetCurrentPerson(self,number):
        self.current_blank.SetLabel(str(number))
    def OnPrev(self, event):
        if not self.IsLoaded:
            return
        cur = self.GetCurrentPerson()
        if cur > 1:
            self.SetCurrentPerson(cur-1)
            self.LoadBlank()
            self.fio.SetValue('')
            self.unit.SetValue('')
    def OnNext(self, event):
        if not self.IsLoaded:
            return
        cur = self.GetCurrentPerson()
        if cur < self.GetMaxPersons():
            self.SetCurrentPerson(cur+1)
            self.LoadBlank()
            self.fio.SetValue('')
            self.unit.SetValue('')
    def OnSave(self, event):
        if not self.IsLoaded:
            return
        unit = self.unit.GetValue()
        fio = self.fio.GetValue()
        if fio == '' or unit == '':
            return
        else:
            font = Font(size=14, name="Times New Roman")
            wndws = []
            for label in self.labels:
                wndws.append(self.FindWindowByName(label).GetLabel())
            if (fio in justdata) and (unit in data.keys()):
                row = justdata.index(fio) + 2
            else:
                cols = [ws.max_row,fio.lower(),unit.lower()]
                row = ws.max_row+1
                for i,c in enumerate(cols,1):
                    cell = ws.cell(row=row,column=i)
                    cell.value = c
                    cell.font = font
            for iter in range(4,10):
                cell = ws.cell(row=row,column=iter)
                cell.font = font
                cell.value = wndws[iter-4].lower()

            wb.save('db.xlsx')
            wx.MessageBox('Успешно сохранено:\n' + unit + ' – ' + fio,'Сохранение')
if __name__ == '__main__':
    app = wx.App()
    frame = Frame(parent=None, id=-1)
    frame.Show()
    app.MainLoop()
